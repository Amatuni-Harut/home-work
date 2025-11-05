from sqlalchemy import create_engine, Column, Integer, String
from sqlalchemy.orm import declarative_base, sessionmaker, Session
from fastapi import FastAPI, HTTPException, Depends, UploadFile, File
from pydantic import BaseModel
from openpyxl import load_workbook
import xlsxwriter
from typing import List
from datetime import datetime
from pathlib import Path

app = FastAPI()

# ------------------------ SQLAlchemy Setup ------------------------
DATABASE_URL = "sqlite:///data_movies.db"
engine = create_engine(DATABASE_URL, echo=False)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)
Base = declarative_base()

# ------------------------ Database Model ------------------------
class Movie(Base):
    __tablename__ = "movies"
    id = Column(Integer, primary_key=True, index=True)
    title = Column(String, nullable=False)
    year = Column(Integer, nullable=False)
    genre = Column(String, nullable=False)
    director = Column(String, nullable=False)

Base.metadata.create_all(bind=engine)

# ------------------------ Pydantic Schema ------------------------
class MovieSchema(BaseModel):
    title: str
    year: int
    genre: str
    director: str

    class Config:
        orm_mode = True

# ------------------------ Database Dependency ------------------------
def get_db():
    db: Session = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# ------------------------ Excel Functions ------------------------
def import_file(fname: str) -> List[dict]:

    wb = load_workbook(fname)
    ws = wb.active
    data: List[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title, year, genre, director = row
        if title and year and genre and director:
            data.append({
                "title": title,
                "year": int(year),
                "genre": genre,
                "director": director
            })
    return data

def export_file(data: List[dict], fname: str):

    workbook = xlsxwriter.Workbook(fname)
    worksheet = workbook.add_worksheet()
    headers = ["Title", "Year", "Genre", "Director"]
    for col, header in enumerate(headers):
        worksheet.write(0, col, header)
    for row, movie in enumerate(data, start=1):
        worksheet.write(row, 0, movie["title"])
        worksheet.write(row, 1, movie["year"])
        worksheet.write(row, 2, movie["genre"])
        worksheet.write(row, 3, movie["director"])
    workbook.close()

def create_excel_backup(db: Session):
 
    movies = db.query(Movie).all()
    data = [{"title": m.title, "year": m.year, "genre": m.genre, "director": m.director} for m in movies]
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    fname = f"movies_backup_{timestamp}.xlsx"
    export_file(data, fname)
    print("Excel backup created:", fname)

# ------------------------ FastAPI CRUD Routes ------------------------
@app.get("/movies", response_model=List[MovieSchema])
def get_all_movies(db: Session = Depends(get_db)):

    return db.query(Movie).all()

@app.get("/movies/{m_id}", response_model=MovieSchema)
def get_movie_by_id(m_id: int, db: Session = Depends(get_db)):
   
    movie = db.query(Movie).filter(Movie.id == m_id).first()
    if not movie:
        raise HTTPException(status_code=404, detail="Movie not found")
    return movie

@app.post("/movies")
def create_movie(movie: MovieSchema, db: Session = Depends(get_db)):

    new_movie = Movie(**movie.dict())
    db.add(new_movie)
    db.commit()
    db.refresh(new_movie)
    create_excel_backup(db)
    return {"msg": "Movie created", "id": new_movie.id}

@app.put("/movies/{m_id}")
def update_movie(m_id: int, movie: MovieSchema, db: Session = Depends(get_db)):
  
    movie_data = db.query(Movie).filter(Movie.id == m_id).first()
    if not movie_data:
        raise HTTPException(status_code=404, detail="Movie ID not found")
    for key, value in movie.dict().items():
        setattr(movie_data, key, value)
    db.commit()
    db.refresh(movie_data)
    create_excel_backup(db)
    return {"msg": "Movie updated"}

@app.delete("/movies/{m_id}")
def delete_movie(m_id: int, db: Session = Depends(get_db)):

    movie_data = db.query(Movie).filter(Movie.id == m_id).first()
    if not movie_data:
        raise HTTPException(status_code=404, detail="Movie ID not found")
    db.delete(movie_data)
    db.commit()
    create_excel_backup(db)
    return {"msg": "Movie deleted"}

# ------------------------ Excel Endpoints ------------------------
@app.post("/import_excel")
def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):

    contents = file.file.read()
    temp_file = "temp.xlsx"
    with open(temp_file, "wb") as f:
        f.write(contents)
    try:
        movies_data = import_file(temp_file)
        for movie_dict in movies_data:
            db.add(Movie(**movie_dict))
        db.commit()
        create_excel_backup(db)
        return {"msg": f"{len(movies_data)} movies imported successfully"}
    except :
        raise HTTPException(status_code=500, detail="Error importing Excel")

@app.get("/export_excel")
def export_excel(db: Session = Depends(get_db)):

    movies = db.query(Movie).all()
    data = [{"title": m.title, "year": m.year, "genre": m.genre, "director": m.director} for m in movies]
    fname = "movies_export.xlsx"
    export_file(data, fname)
    return {"msg": f"{len(data)} movies exported to {fname}"}

# ------------------------ Startup Event ------------------------
@app.on_event("startup")
def startup_import_excel():

    excel_file = "film_data.xlsx"
    db = SessionLocal()
    try:
        if Path(excel_file).exists():
            print(f"Importing file: {excel_file}")
            movies_data = import_file(excel_file)
            if not movies_data:
                print("Excel file is empty.")
                return
            if db.query(Movie).count() == 0:
                for el in movies_data:
                    db.add(Movie(**el))
                db.commit()
                print(f"Imported {len(movies_data)} movies into the database.")
            else:
                print("Table already contains data. Import skipped.")
        else:
            print(f"File {excel_file} not found.")
    except:
        print("Error importing Excel")
    finally:
        db.close()
